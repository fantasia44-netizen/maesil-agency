"""finance_hometax — 홈택스 전자(세금)계산서 엑셀 파서.

maesil-total의 tax_invoice_service.parse_hometax_excel을 이식 (pandas 제거,
openpyxl/xlrd 직접 사용 — agency 백엔드는 pandas 미탑재).

홈택스 엑셀 구조:
  Row 0~4: 메타(사업자정보/합계/제목) — 가변
  헤더 행: '작성일자'와 '승인번호'가 함께 있는 행을 자동 탐지
  세금계산서(과세): '세액' 컬럼 있음 (16번) / 계산서(면세): 없음

컬럼명이 중복(상호·대표자명 등)이라 위치 기반 매핑을 쓴다.
"""
from __future__ import annotations

import logging
from io import BytesIO

logger = logging.getLogger(__name__)


# ── 셀 값 정규화 ─────────────────────────────────────────────────────

def _safe_str(val) -> str:
    if val is None:
        return ""
    s = str(val).strip()
    return "" if s in ("nan", "NaN", "None") else s


def _safe_int(val) -> int:
    s = _safe_str(val)
    if not s:
        return 0
    try:
        return int(float(s.replace(",", "").replace(" ", "")))
    except (ValueError, TypeError):
        return 0


def _normalize_date(val) -> str:
    """다양한 날짜 표현 → YYYY-MM-DD ('' 허용)."""
    s = _safe_str(val).replace(".", "-").replace("/", "-")
    if not s:
        return ""
    if len(s) == 8 and s.isdigit():
        return f"{s[:4]}-{s[4:6]}-{s[6:8]}"
    if len(s) >= 10 and s[4] == "-":
        return s[:10]  # 'YYYY-MM-DD hh:mm:ss'(datetime str) 포함 처리
    return s


def _normalize_corp_num(val) -> str:
    return _safe_str(val).replace("-", "").replace(" ", "")


# ── 파일 → 행 목록 ───────────────────────────────────────────────────

def load_excel_rows(file_bytes: bytes, filename: str) -> list[list]:
    """xlsx(openpyxl) / xls(xlrd) → 행(list) 목록. 값은 원시 타입 그대로."""
    name = (filename or "").lower()
    if name.endswith(".xlsx"):
        from openpyxl import load_workbook
        wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb.worksheets[0]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        wb.close()
        return rows
    if name.endswith(".xls"):
        import xlrd  # xlrd 2.x는 .xls 전용
        wb = xlrd.open_workbook(file_contents=file_bytes)
        sh = wb.sheet_by_index(0)
        out = []
        for i in range(sh.nrows):
            row = []
            for j in range(sh.ncols):
                c = sh.cell(i, j)
                if c.ctype == xlrd.XL_CELL_DATE:
                    try:
                        y, m, d, *_ = xlrd.xldate_as_tuple(c.value, wb.datemode)
                        row.append(f"{y:04d}-{m:02d}-{d:02d}")
                    except Exception:
                        row.append(c.value)
                else:
                    row.append(c.value)
            out.append(row)
        return out
    raise ValueError("지원하지 않는 파일 형식입니다. 홈택스에서 받은 .xlsx 또는 .xls 파일을 올려주세요.")


def _find_header_row(rows: list[list]) -> int | None:
    """'작성일자'+'승인번호'가 함께 있는 행 = 실제 헤더."""
    for i in range(min(10, len(rows))):
        vals = [_safe_str(v) for v in rows[i]]
        if "작성일자" in vals and "승인번호" in vals:
            return i
    return None


# ── 메인 파서 ────────────────────────────────────────────────────────

def parse_hometax_excel(file_bytes: bytes, filename: str,
                        direction: str = "sales",
                        is_tax_exempt: bool = False) -> list[dict]:
    """홈택스 발급/수취 목록 엑셀 → finance_tax_invoices insert용 dict 목록.

    direction: 'sales'(매출) | 'purchase'(매입)
    is_tax_exempt: 면세 계산서 강제 플래그 ('세액' 컬럼 없으면 자동 면세 판정)
    """
    rows = load_excel_rows(file_bytes, filename)

    header_idx = _find_header_row(rows)
    if header_idx is None:
        raise ValueError(
            "홈택스 엑셀 형식을 인식할 수 없습니다. "
            "'작성일자', '승인번호' 컬럼이 포함된 홈택스 표준 양식인지 확인하세요."
        )

    header_vals = [_safe_str(v) for v in rows[header_idx]]
    data_rows = rows[header_idx + 1:]

    # '세액' 컬럼 존재 여부로 과세/면세 자동 판별
    has_tax_col = "세액" in header_vals
    if not has_tax_col:
        is_tax_exempt = True

    # 위치 기반 매핑 (홈택스 표준: 과세 35컬럼 / 면세 33컬럼 — 세액 유무만 차이)
    IDX = {
        "write_date": 0, "invoice_number": 1, "issue_date": 2,
        "supplier_corp_num": 4, "supplier_corp_name": 6, "supplier_ceo_name": 7,
        "buyer_corp_num": 9, "buyer_corp_name": 11, "buyer_ceo_name": 12,
        "total_amount": 14, "supply_cost_total": 15,
        "tax_total": 16 if has_tax_col else None,
        "classification": 17 if has_tax_col else 16,
    }

    def _cell(row, key):
        i = IDX[key]
        return row[i] if (i is not None and i < len(row)) else None

    results: list[dict] = []
    seen: set[str] = set()

    for row in data_rows:
        write_date = _normalize_date(_cell(row, "write_date"))
        if not write_date:
            continue  # 빈 행/합계 행

        issue_date = _normalize_date(_cell(row, "issue_date"))
        invoice_number = _safe_str(_cell(row, "invoice_number"))
        if invoice_number:
            if invoice_number in seen:
                continue  # 파일 내 중복
            seen.add(invoice_number)

        supply = _safe_int(_cell(row, "supply_cost_total"))
        total = _safe_int(_cell(row, "total_amount"))
        tax = 0 if IDX["tax_total"] is None else _safe_int(_cell(row, "tax_total"))

        if supply == 0 and total == 0:
            continue
        if total == 0 and supply != 0:
            total = supply + tax

        if is_tax_exempt:
            tax_type = "면세"
        else:
            classification = _safe_str(_cell(row, "classification"))
            if "영세" in classification:
                tax_type = "영세"
            elif "면세" in classification:
                tax_type = "면세"
            else:
                tax_type = "과세"

        results.append({
            "direction": direction,
            "invoice_number": invoice_number,
            "write_date": write_date or issue_date,
            "issue_date": issue_date or write_date,
            "tax_type": tax_type,
            "supplier_corp_num": _normalize_corp_num(_cell(row, "supplier_corp_num")),
            "supplier_corp_name": _safe_str(_cell(row, "supplier_corp_name")),
            "supplier_ceo_name": _safe_str(_cell(row, "supplier_ceo_name")),
            "buyer_corp_num": _normalize_corp_num(_cell(row, "buyer_corp_num")),
            "buyer_corp_name": _safe_str(_cell(row, "buyer_corp_name")),
            "buyer_ceo_name": _safe_str(_cell(row, "buyer_ceo_name")),
            "supply_cost_total": supply,
            "tax_total": tax,
            "total_amount": total,
            "source": "hometax_excel",
        })

    logger.info("[finance] 홈택스 파싱: %d건 (%s, %s)",
                len(results), direction, "면세" if is_tax_exempt else "과세포함")
    return results
